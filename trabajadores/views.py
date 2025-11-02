# Crear el archivo views.py corregido
import io
from django.contrib.auth.decorators import login_required
from datetime import date, timedelta
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.db.models import Q, Count
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.contrib import messages
from django.core.paginator import Paginator
from django.shortcuts import render
from django.db.models import Prefetch
from .models import (
    TrabajadorPersonal, TrabajadorLaboral, TrabajadorAfiliaciones,
    TrabajadorDotacion, TrabajadorCurso, TrabajadorRol, TipoCurso, TipoDotacion
)
from .forms import (
    TrabajadorPersonalForm, TrabajadorLaboralForm, TrabajadorAfiliacionesForm,
    TrabajadorDotacionForm, TrabajadorCursoForm, TrabajadorRolForm, TipoCursoForm, TipoDotacionForm
)

def home(request):
    """Página principal con menú de iconos"""
    return render(request, 'trabajadores/home.html')

def dashboard_alertas_cursos(request):
    # Traemos todos los cursos con su trabajador y tipo de curso
    cursos = TrabajadorCurso.objects.select_related('id_trabajador', 'tipo_curso').all()

    # Opcional: filtrar por estado si se pasa en GET ?estado=vigente|vencido|proximo_vencer
    estado_filtro = request.GET.get('estado')
    if estado_filtro:
        cursos = [c for c in cursos if c.get_estado_vigencia() == estado_filtro]

    context = {
        'cursos': cursos,
        'estado_filtro': estado_filtro,
    }
    return render(request, 'dashboard_alertas_cursos.html', context)

# ====
# MAPEO COMPLETO DE CAMPOS DISPONIBLES PARA EXPORTACIÓN
# ====
# Formato: "Etiqueta": ("atributo_o_metodo", callable_opcional)
TRABAJADOR_FIELDS = {
    # Identificación
    "Documento": ("id_trabajador", None),
    "Tipo Doc.": ("get_tipo_documento_display", lambda obj: obj.get_tipo_documento_display()),
    "Nombres": ("nombres", None),
    "Apellidos": ("apellidos", None),
    "Fecha expedición doc.": ("fecha_expedicion_doc", None),
    "Lugar expedición": ("lugar_expedicion", None),
    "Nacionalidad": ("nacionalidad", None),
    
    # Datos Personales
    "Fecha nacimiento": ("fecha_nacimiento", None),
    "Lugar nacimiento": ("lugar_nacimiento", None),
    "Género": ("get_genero_display", lambda obj: obj.get_genero_display()),
    "Estado civil": ("get_estado_civil_display", lambda obj: obj.get_estado_civil_display()),
    "Grupo sanguíneo": ("grupo_sanguineo_rh", None),
    "Nombre padre": ("nombres_padre", None),
    "Nombre madre": ("nombres_madre", None),
    "Número de hijos": ("numero_hijos", None),
    
    # Datos Bancarios
    "Cuenta bancaria": ("cuenta_bancaria", None),
    "Tipo cuenta": ("get_tipo_cuenta_display", lambda obj: obj.get_tipo_cuenta_display()),
    "Banco": ("banco", None),
    
    # Contacto
    "Dirección": ("direccion_residencia", None),
    "Ciudad": ("ciudad_residencia", None),
    "Departamento": ("departamento_residencia", None),
    "Celular": ("telefono_celular", None),
    "Correo": ("correo_personal", None),
    "Contacto emergencia": ("nombre_contacto_emergencia", None),
    "Teléfono emergencia": ("numero_contacto_emergencia", None),
    
    # Datos Laborales (último contrato)
    "Cargo (último)": ("cargo_ultimo", None),
    "Salario (último)": ("salario_ultimo", None),
    "Tipo contrato (último)": ("tipo_contrato_ultimo", None),
    "Jornada (último)": ("jornada_ultimo", None),
    "Sede trabajo (último)": ("sede_ultimo", None),
    "Inicio contrato (último)": ("inicio_ultimo", None),
    "Fin contrato (último)": ("fin_ultimo", None),
    
    # Afiliaciones
    "EPS": ("eps_nombre", None),
    "EPS número": ("eps_numero", None),
    "Fondo pensiones": ("pension_nombre", None),
    "Fondo pensiones número": ("pension_numero", None),
    "ARL": ("arl_nombre", None),
    "ARL número": ("arl_numero", None),
    "Caja compensación": ("caja_nombre", None),
    "Caja compensación número": ("caja_numero", None),
    
    # Cursos y Dotaciones
    "Cursos": ("cursos_texto", None),
    "Dotaciones": ("dotaciones_texto", None),
}

# ====
# FUNCIONES AUXILIARES PARA EXPORTACIÓN
# ====

def _query_trabajadores_from_filters(request):
    """
    Construye un queryset de trabajadores aplicando filtros de búsqueda.
    Soporta tanto GET como POST.
    """
    qs = TrabajadorPersonal.objects.all()
    
    # Búsqueda por texto
    q = (request.GET.get('q') or request.POST.get('q') or '').strip()
    if q:
        qs = qs.filter(
            Q(nombres__icontains=q) | 
            Q(apellidos__icontains=q) | 
            Q(id_trabajador__icontains=q)
        )
    
    # Filtros específicos
    for key in ("tipo_documento", "estado_civil", "genero"):
        val = request.GET.get(key) or request.POST.get(key)
        if val:
            qs = qs.filter(**{key: val})
    
    return qs


def _apply_selection(request, queryset):
    """
    Filtra el queryset para incluir solo los IDs seleccionados.
    Espera 'selected_ids' como lista en POST.
    """
    ids = []
    if request.method == 'POST':
        ids = request.POST.getlist('selected_ids')
    
    if ids:
        queryset = queryset.filter(id_trabajador__in=ids)
    
    return queryset


def _annotate_last_laboral(trabajadores):
    """
    Adjunta atributos temporales con datos del último registro laboral,
    para usar en los campos '… (último)' si se seleccionan.
    """
    trabajadores = trabajadores.prefetch_related('registros_laborales', 'afiliaciones', 'cursos', 'dotaciones')
    for t in trabajadores:
        # Último registro laboral
        ult = None
        try:
            ult = sorted(t.registros_laborales.all(), key=lambda x: (x.fecha_inicio_contrato or ''), reverse=True)[0]
        except IndexError:
            pass
        
        t.cargo_ultimo = getattr(ult, 'cargo', '')
        t.salario_ultimo = getattr(ult, 'salario', '')
        t.tipo_contrato_ultimo = getattr(ult, 'tipo_contrato', '')
        t.jornada_ultimo = getattr(ult, 'jornada_laboral', '')
        t.sede_ultimo = getattr(ult, 'sede_trabajo', '')
        t.inicio_ultimo = getattr(ult, 'fecha_inicio_contrato', '')
        t.fin_ultimo = getattr(ult, 'fecha_terminacion_contrato', '')
        
        # Afiliaciones (tomar la primera/única)
        afil = None
        try:
            afil = t.afiliaciones.first()
        except:
            pass
        
        t.eps_nombre = getattr(afil, 'eps_nombre', '')
        t.eps_numero = getattr(afil, 'eps_numero_afiliacion', '')
        t.pension_nombre = getattr(afil, 'fondo_pensiones_nombre', '')
        t.pension_numero = getattr(afil, 'fondo_pensiones_numero_afiliacion', '')
        t.arl_nombre = getattr(afil, 'arl_nombre', '')
        t.arl_numero = getattr(afil, 'arl_numero_nombre', '')
        t.caja_nombre = getattr(afil, 'caja_compensacion_nombre', '')
        t.caja_numero = getattr(afil, 'caja_compensacion_numero_afiliacion', '')
        
        # Cursos (concatenar todos)
        cursos = t.cursos.all()
        if cursos.exists():
            cursos_texto = '; '.join([f"{c.nombre_curso} - {c.institucion} ({c.fecha_inicio_curso})" for c in cursos])
            t.cursos_texto = cursos_texto
        else:
            t.cursos_texto = ''
        
        # Dotaciones (concatenar todas)
        dotaciones = t.dotaciones.all()
        if dotaciones.exists():
            dotaciones_texto = '; '.join([f"{d.tipo_dotacion} - Talla {d.talla} ({d.fecha_entrega})" for d in dotaciones])
            t.dotaciones_texto = dotaciones_texto
        else:
            t.dotaciones_texto = ''
    
    return trabajadores


# ====
# EXPORTACIÓN A EXCEL PERSONALIZADA
# ====

def export_trabajadores_excel_custom(request):
    """
    Exporta trabajadores a Excel con campos personalizados.
    Recibe:
    - selected_ids: Lista de IDs de trabajadores seleccionados
    - fields: Lista de campos a exportar
    """
    # 1) Obtener trabajadores seleccionados
    qs = TrabajadorPersonal.objects.all()
    qs = _apply_selection(request, qs)
    
    if not qs.exists():
        messages.error(request, "No hay trabajadores seleccionados para exportar.")
        return redirect('trabajadores:trabajador_list')
    
    # 2) Campos solicitados
    selected_fields = request.POST.getlist('fields')
    if not selected_fields:
        selected_fields = ["Documento", "Nombres", "Apellidos", "Celular", "Correo"]
    
    # 3) Enriquecer con último laboral/afiliaciones si se requieren
    if any(f.endswith("(último)") for f in selected_fields) or any(f in ["Cursos", "Dotaciones", "EPS", "EPS número", "Fondo pensiones", "Fondo pensiones número", "ARL", "ARL número", "Caja compensación", "Caja compensación número"] for f in selected_fields):
        qs = _annotate_last_laboral(qs)
    
    # 4) Crear Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trabajadores"
    
    # Estilos para encabezados
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Encabezados
    for col_idx, label in enumerate(selected_fields, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Filas de datos
    for row_idx, obj in enumerate(qs.order_by('apellidos', 'nombres'), start=2):
        for col_idx, label in enumerate(selected_fields, start=1):
            attr, call = TRABAJADOR_FIELDS.get(label, (None, None))
            if call:
                val = call(obj)
            elif attr:
                val = getattr(obj, attr, '')
            else:
                val = ''
            ws.cell(row=row_idx, column=col_idx, value=val)
    
    # Ajustar ancho de columnas
    for col_idx, header in enumerate(selected_fields, start=1):
        max_len = len(header)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            v = row[0].value
            max_len = max(max_len, len(str(v)) if v is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)
    
    # Respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=trabajadores_seleccionados.xlsx'
    wb.save(response)
    return response


# ====
# EXPORTACIÓN A PDF PERSONALIZADA
# ====

def render_to_pdf(template_src, context):
    """
    Renderiza un template HTML a PDF usando xhtml2pdf (pisa).
    Devuelve bytes del PDF o None si hay error.
    """
    template = get_template(template_src)
    html = template.render(context)
    result = io.BytesIO()
    pdf = pisa.pisaDocument(
        io.BytesIO(html.encode("UTF-8")),
        dest=result,
        encoding='UTF-8'
    )
    if not pdf.err:
        return result.getvalue()
    return None


def export_trabajadores_pdf_custom(request):
    """
    Exporta un reporte PDF con la lista de trabajadores filtrados/seleccionados.
    Similar a export_trabajadores_excel_custom pero genera un PDF.
    """
    # 1) Trabajadores por filtros y/o selección de checkboxes
    qs = _query_trabajadores_from_filters(request)
    qs = _apply_selection(request, qs)
    
    # 2) Campos solicitados
    selected_fields = request.POST.getlist('fields') or request.GET.getlist('fields')
    if not selected_fields:
        selected_fields = ["Documento", "Nombres", "Apellidos", "Celular", "Correo"]
    
    # 3) Enriquecer con último laboral y afiliaciones si se requieren
    if any(f.endswith("(último)") for f in selected_fields) or any(f in ["Cursos", "Dotaciones", "EPS", "EPS número", "Fondo pensiones", "Fondo pensiones número", "ARL", "ARL número", "Caja compensación", "Caja compensación número"] for f in selected_fields):
        qs = _annotate_last_laboral(qs)
    
    # 4) Preparar datos para el template
    trabajadores_data = []
    for obj in qs.order_by('apellidos', 'nombres'):
        row = {}
        for label in selected_fields:
            if label in TRABAJADOR_FIELDS:
                attr, call = TRABAJADOR_FIELDS[label]
                if call:
                    val = call(obj)
                else:
                    val = getattr(obj, attr, '')
                row[label] = val if val else '-'
            else:
                row[label] = '-'
        trabajadores_data.append(row)
    
    # 5) Contexto para el template
    context = {
        'trabajadores': trabajadores_data,
        'campos': selected_fields,
        'total': len(trabajadores_data),
    }
    
    # 6) Generar PDF
    pdf_bytes = render_to_pdf('trabajadores/report_trabajadores_custom.html', context)
    if not pdf_bytes:
        return HttpResponse("Error generando PDF", status=500)
    
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="trabajadores_reporte.pdf"'
    return response


def export_trabajador_pdf(request, id_trabajador):
    """
    Exporta un trabajador individual a PDF con todos sus datos relacionados.
    """
    trabajador = get_object_or_404(TrabajadorPersonal, pk=id_trabajador)
    context = {
        'trabajador': trabajador,
        'laborales': trabajador.registros_laborales.all().order_by('-fecha_inicio_contrato'),
        'afiliaciones': trabajador.afiliaciones.all(),
        'dotaciones': trabajador.dotaciones.all().order_by('-fecha_entrega'),
        'cursos': trabajador.cursos.all().order_by('-fecha_inicio_curso'),
        'roles': trabajador.roles_sistema.all(),
    }
    
    pdf_bytes = render_to_pdf('trabajadores/report_trabajador_detail.html', context)
    if not pdf_bytes:
        return HttpResponse("Error generando PDF", status=500)
    
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="trabajador_{trabajador.id_trabajador}.pdf"'
    return response


# ====
# VISTAS CRUD - TrabajadorPersonal
# ====

def trabajador_list(request):
    """
    Vista de listado de trabajadores con búsqueda, filtros y paginación.
    
    Parámetros GET:
    - q: Búsqueda por nombre, apellido o documento
    - tipo_documento: Filtro por tipo de documento
    - estado_civil: Filtro por estado civil
    - genero: Filtro por género
    - ordenar: Campo por el cual ordenar (apellidos, nombres, id_trabajador)
    """
    
    # Obtener todos los trabajadores
    trabajadores = TrabajadorPersonal.objects.all()
    
    # ==== BÚSQUEDA ====
    query = request.GET.get('q', '').strip()
    if query:
        trabajadores = trabajadores.filter(
            Q(nombres__icontains=query) |
            Q(apellidos__icontains=query) |
            Q(id_trabajador__icontains=query)
        )
    
    # ==== FILTROS ====
    tipo_documento = request.GET.get('tipo_documento', '')
    if tipo_documento:
        trabajadores = trabajadores.filter(tipo_documento=tipo_documento)
    
    estado_civil = request.GET.get('estado_civil', '')
    if estado_civil:
        trabajadores = trabajadores.filter(estado_civil=estado_civil)
    
    genero = request.GET.get('genero', '')
    if genero:
        trabajadores = trabajadores.filter(genero=genero)
    
    # ==== ORDENAMIENTO ====
    ordenar = request.GET.get('ordenar', 'apellidos')
    campos_validos = ['apellidos', 'nombres', 'id_trabajador', '-apellidos', '-nombres', '-id_trabajador']
    if ordenar in campos_validos:
        trabajadores = trabajadores.order_by(ordenar, 'nombres')
    else:
        trabajadores = trabajadores.order_by('apellidos', 'nombres')
    
    # ==== PAGINACIÓN ====
    paginator = Paginator(trabajadores, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # ==== CONTEXTO ====
    context = {
        'trabajadores': page_obj,
        'page_obj': page_obj,
        'query': query,
        'tipo_documento': tipo_documento,
        'estado_civil': estado_civil,
        'genero': genero,
        'ordenar': ordenar,
        'total_trabajadores': paginator.count,
        'tipos_documento': TrabajadorPersonal._meta.get_field('tipo_documento').choices,
        'estados_civiles': TrabajadorPersonal._meta.get_field('estado_civil').choices,
        'generos': TrabajadorPersonal._meta.get_field('genero').choices,
    }
    
    return render(request, 'trabajadores/trabajador_list.html', context)


def trabajador_detail(request, id_trabajador):
    trabajador = get_object_or_404(TrabajadorPersonal, pk=id_trabajador)
    laborales = trabajador.registros_laborales.all().order_by('-fecha_inicio_contrato')
    afiliaciones = trabajador.afiliaciones.all()
    dotaciones = trabajador.dotaciones.all().order_by('-fecha_entrega')
    cursos = trabajador.cursos.all().order_by('-fecha_inicio_curso')
    roles = trabajador.roles_sistema.all()
    context = {
        'trabajador': trabajador,
        'laborales': laborales,
        'afiliaciones': afiliaciones,
        'dotaciones': dotaciones,
        'cursos': cursos,
        'roles': roles,
    }
    return render(request, 'trabajadores/trabajador_detail.html', context)


def trabajador_create(request):
    if request.method == 'POST':
        form = TrabajadorPersonalForm(request.POST)
        if form.is_valid():
            obj = form.save()
            messages.success(request, "Trabajador creado correctamente.")
            return redirect('trabajadores:trabajador_detail', id_trabajador=obj.id_trabajador)
    else:
        form = TrabajadorPersonalForm()
    return render(request, 'trabajadores/trabajador_form.html', {'form': form, 'title': 'Nuevo Trabajador'})


def trabajador_update(request, id_trabajador):
    trabajador = get_object_or_404(TrabajadorPersonal, pk=id_trabajador)
    if request.method == 'POST':
        form = TrabajadorPersonalForm(request.POST, instance=trabajador)
        if form.is_valid():
            form.save()
            nuevo_id = form.instance.id_trabajador
            messages.success(request, "Trabajador actualizado correctamente.")
            return redirect('trabajadores:trabajador_detail', id_trabajador=nuevo_id)
    else:
        form = TrabajadorPersonalForm(instance=trabajador)
    return render(request, 'trabajadores/trabajador_form.html', {'form': form, 'title': 'Editar Trabajador'})


def trabajador_delete(request, id_trabajador):
    trabajador = get_object_or_404(TrabajadorPersonal, pk=id_trabajador)
    if request.method == 'POST':
        trabajador.delete()
        messages.success(request, "Trabajador eliminado.")
        return redirect('trabajadores:trabajador_list')
    return render(request, 'trabajadores/confirm_delete.html', {
        'obj': trabajador,
        'cancel_url': reverse('trabajadores:trabajador_detail', kwargs={'id_trabajador': id_trabajador})
    })


# ====
# VISTAS CRUD - Laboral
# ====

def laboral_create(request, id_trabajador):
    trabajador = get_object_or_404(TrabajadorPersonal, pk=id_trabajador)
    
    if request.method == 'POST':
        form = TrabajadorLaboralForm(request.POST)
        if 'id_trabajador' in form.fields:
            form.fields.pop('id_trabajador')
        
        if form.is_valid():
            obj = form.save(commit=False)
            obj.id_trabajador = trabajador
            obj.save()
            messages.success(request, "Registro laboral creado.")
            return redirect('trabajadores:trabajador_detail', id_trabajador=id_trabajador)
    else:
        form = TrabajadorLaboralForm()
        form.fields.pop('id_trabajador')
    
    return render(request, 'trabajadores/relacionado_form.html', {
        'form': form,
        'title': 'Nuevo Registro Laboral',
        'trabajador': trabajador
    })


def laboral_update(request, id_laboral):
    laboral = get_object_or_404(TrabajadorLaboral, pk=id_laboral)
    trabajador = laboral.id_trabajador
    
    if request.method == 'POST':
        form = TrabajadorLaboralForm(request.POST, instance=laboral)
        if 'id_trabajador' in form.fields:
            form.fields.pop('id_trabajador')
        
        if form.is_valid():
            form.save()
            messages.success(request, "Registro laboral actualizado.")
            return redirect('trabajadores:trabajador_detail', id_trabajador=trabajador.id_trabajador)
    else:
        form = TrabajadorLaboralForm(instance=laboral)
        form.fields.pop('id_trabajador')
    
    return render(request, 'trabajadores/relacionado_form.html', {
        'form': form,
        'title': 'Editar Registro Laboral',
        'trabajador': trabajador
    })


def laboral_delete(request, id_laboral):
    laboral = get_object_or_404(TrabajadorLaboral, pk=id_laboral)
    trabajador_id = laboral.id_trabajador.id_trabajador
    if request.method == 'POST':
        laboral.delete()
        messages.success(request, "Registro laboral eliminado.")
        return redirect('trabajadores:trabajador_detail', id_trabajador=trabajador_id)
    return render(request, 'trabajadores/confirm_delete.html', {
        'obj': laboral,
        'cancel_url': reverse('trabajadores:trabajador_detail', kwargs={'id_trabajador': trabajador_id})
    })


# ====
# VISTAS CRUD - Afiliaciones
# ====

def afiliacion_create(request, id_trabajador):
    trabajador = get_object_or_404(TrabajadorPersonal, pk=id_trabajador)
    
    if request.method == 'POST':
        form = TrabajadorAfiliacionesForm(request.POST)
        if 'id_trabajador' in form.fields:
            form.fields.pop('id_trabajador')
        
        if form.is_valid():
            obj = form.save(commit=False)
            obj.id_trabajador = trabajador
            obj.save()
            messages.success(request, "Afiliación creada.")
            return redirect('trabajadores:trabajador_detail', id_trabajador=id_trabajador)
    else:
        form = TrabajadorAfiliacionesForm()
        form.fields.pop('id_trabajador')
    
    return render(request, 'trabajadores/relacionado_form.html', {
        'form': form,
        'title': 'Nueva Afiliación',
        'trabajador': trabajador
    })


def afiliacion_update(request, id_afiliacion):
    afiliacion = get_object_or_404(TrabajadorAfiliaciones, pk=id_afiliacion)
    trabajador = afiliacion.id_trabajador
    
    if request.method == 'POST':
        form = TrabajadorAfiliacionesForm(request.POST, instance=afiliacion)
        if 'id_trabajador' in form.fields:
            form.fields.pop('id_trabajador')
        
        if form.is_valid():
            form.save()
            messages.success(request, "Afiliación actualizada.")
            return redirect('trabajadores:trabajador_detail', id_trabajador=trabajador.id_trabajador)
    else:
        form = TrabajadorAfiliacionesForm(instance=afiliacion)
        form.fields.pop('id_trabajador')
    
    return render(request, 'trabajadores/relacionado_form.html', {
        'form': form,
        'title': 'Editar Afiliación',
        'trabajador': trabajador
    })


def afiliacion_delete(request, id_afiliacion):
    afiliacion = get_object_or_404(TrabajadorAfiliaciones, pk=id_afiliacion)
    trabajador_id = afiliacion.id_trabajador.id_trabajador
    if request.method == 'POST':
        afiliacion.delete()
        messages.success(request, "Afiliación eliminada.")
        return redirect('trabajadores:trabajador_detail', id_trabajador=trabajador_id)
    return render(request, 'trabajadores/confirm_delete.html', {
        'obj': afiliacion,
        'cancel_url': reverse('trabajadores:trabajador_detail', kwargs={'id_trabajador': trabajador_id})
    })


# ====
# VISTAS CRUD - Dotación
# ====

def dotacion_create(request, id_trabajador):
    """Vista para crear una nueva dotación"""
    trabajador = get_object_or_404(TrabajadorPersonal, id_trabajador=id_trabajador)
    tipos_dotacion = TipoDotacion.objects.filter(activo=True)
    
    if request.method == 'POST':
        print("=== DEBUG DOTACION CREATE ===")
        print(f"POST data: {request.POST}")
        
        form = TrabajadorDotacionForm(request.POST)
        
        print(f"Form is valid: {form.is_valid()}")
        
        if form.is_valid():
            print(f"Cleaned data: {form.cleaned_data}")
            
            # Guardar sin commit para poder asignar el trabajador
            dotacion = form.save(commit=False)
            
            # ⭐ CRÍTICO: Asignar el trabajador
            dotacion.id_trabajador = trabajador
            
            # ⭐ CRÍTICO: Asegurar que tipo_dotacion esté sincronizado
            if dotacion.tipo_dotacion_catalogo:
                dotacion.tipo_dotacion = dotacion.tipo_dotacion_catalogo.nombre_tipo_dotacion
            
            # ⭐ CRÍTICO: Estado por defecto
            if not dotacion.estado:
                dotacion.estado = 'ACTIVO'
            
            print(f"Dotacion antes de guardar: {dotacion}")
            print(f"  - id_trabajador: {dotacion.id_trabajador}")
            print(f"  - tipo_dotacion_catalogo: {dotacion.tipo_dotacion_catalogo}")
            print(f"  - tipo_dotacion: {dotacion.tipo_dotacion}")
            print(f"  - talla: {dotacion.talla}")
            print(f"  - estado: {dotacion.estado}")
            
            try:
                dotacion.save()
                print(f"✅ Dotacion guardada con ID: {dotacion.id_dotacion}")
                
                messages.success(
                    request, 
                    f'Dotación {dotacion.tipo_dotacion} (talla {dotacion.talla}) registrada exitosamente.'
                )
                return redirect('trabajadores:trabajador_detail', id_trabajador=id_trabajador)
                
            except Exception as e:
                print(f"❌ Error al guardar: {e}")
                messages.error(request, f'Error al guardar: {str(e)}')
        else:
            print(f"❌ Form errors: {form.errors}")
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = TrabajadorDotacionForm()
    
    context = {
        'form': form,
        'trabajador': trabajador,
        'tipos_dotacion': tipos_dotacion,
    }
    
    return render(request, 'trabajadores/dotacion_form.html', context)



# =====================================================
# TAMBIÉN ASEGÚRATE DE TENER ESTA VISTA PARA EDITAR
# =====================================================

def dotacion_update(request, id_dotacion):
    """Vista para editar una dotación existente"""
    dotacion = get_object_or_404(TrabajadorDotacion, id_dotacion=id_dotacion)
    trabajador = dotacion.id_trabajador
    tipos_dotacion = TipoDotacion.objects.filter(activo=True)

    if request.method == 'POST':
        form = TrabajadorDotacionForm(request.POST, instance=dotacion)
        if form.is_valid():
            dotacion = form.save(commit=False)
            # sincronizar nombre legible
            if dotacion.tipo_dotacion_catalogo:
                dotacion.tipo_dotacion = dotacion.tipo_dotacion_catalogo.nombre_tipo_dotacion
            dotacion.save()
            messages.success(request, 'Dotación actualizada exitosamente.')
            return redirect('trabajadores:trabajador_detail', id_trabajador=trabajador.id_trabajador)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = TrabajadorDotacionForm(instance=dotacion)

    context = {
        'form': form,
        'trabajador': trabajador,
        'tipos_dotacion': tipos_dotacion,
        'dotacion': dotacion,
    }
    return render(request, 'trabajadores/dotacion_form.html', context)

def dotacion_delete(request, id_dotacion):
    """
    Eliminar una dotación
    """
    dotacion = get_object_or_404(TrabajadorDotacion, id_dotacion=id_dotacion)
    trabajador = dotacion.id_trabajador
    
    if request.method == 'POST':
        tipo_dotacion = dotacion.tipo_dotacion
        dotacion.delete()
        messages.success(request, f'Dotación "{tipo_dotacion}" eliminada.')
        return redirect('trabajadores:trabajador_detail', id_trabajador=trabajador.id_trabajador)
    
    context = {
        'objeto': dotacion,
        'tipo': 'Dotación',
        'detalle': f'{dotacion.tipo_dotacion} - Talla {dotacion.talla}',
        'url_cancelar': 'trabajadores:trabajador_detail',
        'id_trabajador': trabajador.id_trabajador
    }
    return render(request, 'trabajadores/confirm_delete.html', context)



# ====
# VISTAS CRUD - Cursos
# ====

def curso_create(request, id_trabajador):
    trabajador = get_object_or_404(TrabajadorPersonal, pk=id_trabajador)

    if request.method == 'POST':
        form = TrabajadorCursoForm(request.POST)
        # no existe id_trabajador en form; lo asignamos abajo
        if form.is_valid():
            obj = form.save(commit=False)
            obj.id_trabajador = trabajador
            # Opcional: si no escriben nombre_curso, puedes copiar el del tipo
            if not obj.nombre_curso and obj.tipo_curso:
                obj.nombre_curso = obj.tipo_curso.nombre_tipo_curso
            obj.save()
            messages.success(request, "Curso registrado.")
            return redirect('trabajadores:trabajador_detail', id_trabajador=id_trabajador)
    else:
        form = TrabajadorCursoForm()
        # En create, puedes limitar es_renovacion_de a cursos del trabajador
        form.fields['es_renovacion_de'].queryset = TrabajadorCurso.objects.filter(id_trabajador=trabajador)

    return render(request, 'trabajadores/curso_form.html', {
        'form': form,
        'title': 'Nuevo Curso',
        'trabajador': trabajador
    })


def curso_update(request, id_curso):
    curso = get_object_or_404(TrabajadorCurso, pk=id_curso)
    trabajador = curso.id_trabajador

    if request.method == 'POST':
        form = TrabajadorCursoForm(request.POST, instance=curso)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.save()
            messages.success(request, "Curso actualizado.")
            return redirect('trabajadores:trabajador_detail', id_trabajador=trabajador.id_trabajador)
    else:
        form = TrabajadorCursoForm(instance=curso)
        form.fields['es_renovacion_de'].queryset = TrabajadorCurso.objects.filter(
            id_trabajador=trabajador
        ).exclude(id_curso=curso.id_curso)

    return render(request, 'trabajadores/curso_form.html', {
        'form': form,
        'title': 'Editar Curso',
        'trabajador': trabajador
    })

def curso_delete(request, id_curso):
    curso = get_object_or_404(TrabajadorCurso, pk=id_curso)
    trabajador_id = curso.id_trabajador.id_trabajador
    if request.method == 'POST':
        curso.delete()
        messages.success(request, "Curso eliminado.")
        return redirect('trabajadores:trabajador_detail', id_trabajador=trabajador_id)
    return render(request, 'trabajadores/confirm_delete.html', {
        'obj': curso,
        'cancel_url': reverse('trabajadores:trabajador_detail', kwargs={'id_trabajador': trabajador_id})
    })


# ====
# VISTAS CRUD - Roles
# ====

def rol_create(request, id_trabajador):
    trabajador = get_object_or_404(TrabajadorPersonal, pk=id_trabajador)
    
    if request.method == 'POST':
        form = TrabajadorRolForm(request.POST)
        if 'id_trabajador' in form.fields:
            form.fields.pop('id_trabajador')
        
        if form.is_valid():
            obj = form.save(commit=False)
            obj.id_trabajador = trabajador
            obj.save()
            messages.success(request, "Rol creado.")
            return redirect('trabajadores:trabajador_detail', id_trabajador=id_trabajador)
    else:
        form = TrabajadorRolForm()
        form.fields.pop('id_trabajador')
    
    return render(request, 'trabajadores/relacionado_form.html', {
        'form': form,
        'title': 'Nuevo Rol',
        'trabajador': trabajador
    })


def rol_update(request, id_rol_sistema):
    rol = get_object_or_404(TrabajadorRol, pk=id_rol_sistema)
    trabajador = rol.id_trabajador
    
    if request.method == 'POST':
        form = TrabajadorRolForm(request.POST, instance=rol)
        if 'id_trabajador' in form.fields:
            form.fields.pop('id_trabajador')
        
        if form.is_valid():
            form.save()
            messages.success(request, "Rol actualizado.")
            return redirect('trabajadores:trabajador_detail', id_trabajador=trabajador.id_trabajador)
    else:
        form = TrabajadorRolForm(instance=rol)
        form.fields.pop('id_trabajador')
    
    return render(request, 'trabajadores/relacionado_form.html', {
        'form': form,
        'title': 'Editar Rol',
        'trabajador': trabajador
    })


def rol_delete(request, id_rol_sistema):
    rol = get_object_or_404(TrabajadorRol, pk=id_rol_sistema)
    trabajador_id = rol.id_trabajador.id_trabajador
    if request.method == 'POST':
        rol.delete()
        messages.success(request, "Rol eliminado.")
        return redirect('trabajadores:trabajador_detail', id_trabajador=trabajador_id)
    return render(request, 'trabajadores/confirm_delete.html', {
        'obj': rol,
        'cancel_url': reverse('trabajadores:trabajador_detail', kwargs={'id_trabajador': trabajador_id})
    })
def dashboard_alertas_cursos(request):
    """
    Dashboard completo de alertas de cursos
    Muestra: Vencidos, Próximos a vencer, Vigentes
    """
    # Traemos todos los cursos con su trabajador y tipo de curso
    cursos = TrabajadorCurso.objects.select_related(
        'id_trabajador', 'tipo_curso'
    ).filter(tipo_curso__isnull=False).order_by('fecha_fin_curso')

    # Filtrar por estado si se pasa en GET
    estado_filtro = request.GET.get('estado')
    
    # Calcular estadísticas
    cursos_evaluados = []
    for curso in cursos:
        estado = curso.get_estado_vigencia()
        if not estado_filtro or estado == estado_filtro:
            cursos_evaluados.append({
                'curso': curso,
                'estado': estado,
                'dias_restantes': curso.dias_para_vencer(),
                'fecha_vencimiento': curso.calcular_fecha_vencimiento()
            })
    
    # Estadísticas generales
    total_cursos = len(cursos_evaluados)
    vigentes = sum(1 for c in cursos_evaluados if c['estado'] == 'vigente')
    proximos_vencer = sum(1 for c in cursos_evaluados if c['estado'] == 'proximo_vencer')
    vencidos = sum(1 for c in cursos_evaluados if c['estado'] == 'vencido')

    context = {
        'cursos_evaluados': cursos_evaluados,
        'estado_filtro': estado_filtro,
        'total_cursos': total_cursos,
        'vigentes': vigentes,
        'proximos_vencer': proximos_vencer,
        'vencidos': vencidos,
    }
    
    return render(request, 'trabajadores/dashboard_alertas_cursos.html', context)


def dashboard_alertas_dotaciones(request):
    """
    Dashboard completo de alertas de dotaciones
    Similar al de cursos pero para dotaciones
    """
    # Traemos dotaciones activas con tipo de dotación
    dotaciones = TrabajadorDotacion.objects.select_related(
        'id_trabajador', 'tipo_dotacion_catalogo'
    ).filter(
        estado='ACTIVO',
        tipo_dotacion_catalogo__isnull=False
    ).order_by('fecha_vencimiento')

    # Filtrar por estado si se pasa en GET
    estado_filtro = request.GET.get('estado')
    
    # Calcular estadísticas
    dotaciones_evaluadas = []
    for dotacion in dotaciones:
        estado = dotacion.get_estado_vigencia()
        if not estado_filtro or estado == estado_filtro:
            dotaciones_evaluadas.append({
                'dotacion': dotacion,
                'estado': estado,
                'dias_restantes': dotacion.dias_para_vencer(),
                'fecha_vencimiento': dotacion.calcular_fecha_vencimiento()
            })
    
    # Estadísticas generales
    total_dotaciones = len(dotaciones_evaluadas)
    vigentes = sum(1 for d in dotaciones_evaluadas if d['estado'] == 'vigente')
    proximos_vencer = sum(1 for d in dotaciones_evaluadas if d['estado'] == 'proximo_vencer')
    vencidos = sum(1 for d in dotaciones_evaluadas if d['estado'] == 'vencido')

    context = {
        'dotaciones_evaluadas': dotaciones_evaluadas,
        'estado_filtro': estado_filtro,
        'total_dotaciones': total_dotaciones,
        'vigentes': vigentes,
        'proximos_vencer': proximos_vencer,
        'vencidos': vencidos,
    }
    
    return render(request, 'trabajadores/dashboard_alertas_dotaciones.html', context)


def dashboard_general(request):
    """
    Dashboard general que muestra resumen de cursos y dotaciones
    """
    # Obtener trabajadores con información de cursos y dotaciones
    trabajadores = TrabajadorPersonal.objects.prefetch_related(
        Prefetch('cursos', queryset=TrabajadorCurso.objects.filter(tipo_curso__isnull=False)),
        Prefetch('dotaciones', queryset=TrabajadorDotacion.objects.filter(estado='ACTIVO'))
    )

    # Calcular alertas por trabajador
    trabajadores_con_alertas = []
    
    for trabajador in trabajadores:
        cursos_vencidos = trabajador.cursos_vencidos()
        cursos_proximos = trabajador.cursos_proximos_vencer()
        dotaciones_proximas = trabajador.dotaciones_proximas_vencer()
        
        # Solo incluir si tiene alertas
        if cursos_vencidos or cursos_proximos or dotaciones_proximas:
            trabajadores_con_alertas.append({
                'trabajador': trabajador,
                'cursos_vencidos': cursos_vencidos,
                'cursos_proximos': cursos_proximos,
                'dotaciones_proximas': dotaciones_proximas,
                'total_alertas': len(cursos_vencidos) + len(cursos_proximos) + len(dotaciones_proximas)
            })
    
    # Ordenar por cantidad de alertas (descendente)
    trabajadores_con_alertas.sort(key=lambda x: x['total_alertas'], reverse=True)

    # Estadísticas generales
    context = {
        'trabajadores_con_alertas': trabajadores_con_alertas,
        'total_trabajadores_con_alertas': len(trabajadores_con_alertas),
        'total_alertas': sum(t['total_alertas'] for t in trabajadores_con_alertas),
    }
    
    return render(request, 'trabajadores/dashboard_general.html', context)


def reporte_trabajadores_sin_documentacion(request):
    """
    Reporte de trabajadores que NO tienen documentación completa
    (cursos vencidos o sin dotación activa)
    """
    trabajadores = TrabajadorPersonal.objects.prefetch_related(
        'cursos__tipo_curso',
        'dotaciones'
    )
    
    trabajadores_incompletos = []
    
    for trabajador in trabajadores:
        # Verificar cursos vencidos
        cursos_vencidos = [c for c in trabajador.cursos.all() 
                          if c.tipo_curso and c.get_estado_vigencia() == 'vencido']
        
        # Verificar si tiene dotación activa
        tiene_dotacion = trabajador.dotaciones.filter(estado='ACTIVO').exists()
        
        if cursos_vencidos or not tiene_dotacion:
            trabajadores_incompletos.append({
                'trabajador': trabajador,
                'cursos_vencidos': cursos_vencidos,
                'sin_dotacion': not tiene_dotacion,
            })
    
    context = {
        'trabajadores_incompletos': trabajadores_incompletos,
        'total': len(trabajadores_incompletos),
    }
    
    return render(request, 'trabajadores/reporte_documentacion_incompleta.html', context)
# =====================================================
# VISTAS PARA GESTIONAR TIPOS DE CURSOS
# =====================================================

@login_required
def tipo_curso_list(request):
    """Listado de tipos de cursos"""
    tipos = TipoCurso.objects.all().order_by('nombre_tipo_curso')
    context = {
        'tipos': tipos,
        'total_activos': tipos.filter(activo=True).count(),
        'total_inactivos': tipos.filter(activo=False).count(),
    }
    return render(request, 'trabajadores/tipo_curso_list.html', context)


@login_required
def tipo_curso_create(request):
    """Crear nuevo tipo de curso"""
    if request.method == 'POST':
        form = TipoCursoForm(request.POST)
        if form.is_valid():
            tipo = form.save()
            messages.success(request, f'Tipo de curso "{tipo.nombre_tipo_curso}" creado exitosamente.')
            return redirect('trabajadores:tipo_curso_list')
    else:
        form = TipoCursoForm()
    
    return render(request, 'trabajadores/tipo_curso_form.html', {
        'form': form,
        'title': 'Crear Tipo de Curso'
    })


@login_required
def tipo_curso_update(request, id_tipo_curso):
    """Editar tipo de curso existente"""
    tipo = get_object_or_404(TipoCurso, id_tipo_curso=id_tipo_curso)
    
    if request.method == 'POST':
        form = TipoCursoForm(request.POST, instance=tipo)
        if form.is_valid():
            tipo = form.save()
            messages.success(request, f'Tipo de curso "{tipo.nombre_tipo_curso}" actualizado exitosamente.')
            return redirect('trabajadores:tipo_curso_list')
    else:
        form = TipoCursoForm(instance=tipo)
    
    return render(request, 'trabajadores/tipo_curso_form.html', {
        'form': form,
        'title': 'Editar Tipo de Curso',
        'tipo': tipo
    })


@login_required
def tipo_curso_delete(request, id_tipo_curso):
    """Eliminar tipo de curso"""
    tipo = get_object_or_404(TipoCurso, id_tipo_curso=id_tipo_curso)
    
    if request.method == 'POST':
        nombre = tipo.nombre_tipo_curso
        tipo.delete()
        messages.success(request, f'Tipo de curso "{nombre}" eliminado exitosamente.')
        return redirect('trabajadores:tipo_curso_list')
    
    # Verificar si tiene cursos asociados
    cursos_asociados = tipo.cursos_realizados.count()
    
    return render(request, 'trabajadores/tipo_curso_confirm_delete.html', {
        'tipo': tipo,
        'cursos_asociados': cursos_asociados
    })


# =====================================================
# VISTAS PARA GESTIONAR TIPOS DE DOTACIÓN
# =====================================================

@login_required
def tipo_dotacion_list(request):
    """Listado de tipos de dotaciones"""
    tipos = TipoDotacion.objects.all().order_by('nombre_tipo_dotacion')
    context = {
        'tipos': tipos,
        'total_activos': tipos.filter(activo=True).count(),
        'total_inactivos': tipos.filter(activo=False).count(),
    }
    return render(request, 'trabajadores/tipo_dotacion_list.html', context)


@login_required
def tipo_dotacion_create(request):
    """Crear nuevo tipo de dotación"""
    if request.method == 'POST':
        form = TipoDotacionForm(request.POST)
        if form.is_valid():
            tipo = form.save()
            messages.success(request, f'Tipo de dotación "{tipo.nombre_tipo_dotacion}" creado exitosamente.')
            return redirect('trabajadores:tipo_dotacion_list')
    else:
        form = TipoDotacionForm()
    
    return render(request, 'trabajadores/tipo_dotacion_form.html', {
        'form': form,
        'title': 'Crear Tipo de Dotación'
    })


@login_required
def tipo_dotacion_update(request, id_tipo_dotacion):
    """Editar tipo de dotación existente"""
    tipo = get_object_or_404(TipoDotacion, id_tipo_dotacion=id_tipo_dotacion)
    
    if request.method == 'POST':
        form = TipoDotacionForm(request.POST, instance=tipo)
        if form.is_valid():
            tipo = form.save()
            messages.success(request, f'Tipo de dotación "{tipo.nombre_tipo_dotacion}" actualizado exitosamente.')
            return redirect('trabajadores:tipo_dotacion_list')
    else:
        form = TipoDotacionForm(instance=tipo)
    
    return render(request, 'trabajadores/tipo_dotacion_form.html', {
        'form': form,
        'title': 'Editar Tipo de Dotación',
        'tipo': tipo
    })


@login_required
def tipo_dotacion_delete(request, id_tipo_dotacion):
    """Eliminar tipo de dotación"""
    tipo = get_object_or_404(TipoDotacion, id_tipo_dotacion=id_tipo_dotacion)
    
    if request.method == 'POST':
        nombre = tipo.nombre_tipo_dotacion
        tipo.delete()
        messages.success(request, f'Tipo de dotación "{nombre}" eliminado exitosamente.')
        return redirect('trabajadores:tipo_dotacion_list')
    
    # Verificar si tiene dotaciones asociadas
    dotaciones_asociadas = tipo.dotaciones_entregadas.count()
    
    return render(request, 'trabajadores/tipo_dotacion_confirm_delete.html', {
        'tipo': tipo,
        'dotaciones_asociadas': dotaciones_asociadas
    })
def home(request):
    """Página de inicio con menú de módulos"""
    return render(request, 'trabajadores/home.html')